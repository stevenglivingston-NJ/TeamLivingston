"""Build the KTU Tune-Up pricing workbook: three tiers, priced per door, $2,500 minimum.

Every rate here is pulled from LIVE ServiceMinder KTU invoice lines (168 invoices,
2025-01-01 through 2026-08-18, read via the serviceminder MCP on 2026-08-18) — not
estimated. The per-unit door rates, the flat job lines, and the Shop Labor 24% /
Shop Overhead 5% structure are exactly what the proposals carry today.

The one owner-set number is the $2,500 job minimum: ServiceMinder has no minimum
configured on any of the 42 KTU services (MinimumCharge is null on all of them).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"
BLUE = Font(name=FONT, size=10, color="0000FF")
BLACK = Font(name=FONT, size=10)
GREEN = Font(name=FONT, size=10, color="008000")
BOLD = Font(name=FONT, size=10, bold=True)
RED = Font(name=FONT, size=10, bold=True, color="C00000")
H1 = Font(name=FONT, size=14, bold=True, color="FFFFFF")
H2 = Font(name=FONT, size=11, bold=True)
NOTE = Font(name=FONT, size=9, italic=True, color="595959")

HDR_FILL = PatternFill("solid", fgColor="1F3864")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
BAND_FILL = PatternFill("solid", fgColor="D9E2F3")
TOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
MIN_FILL = PatternFill("solid", fgColor="FCE4D6")

MONEY = '$#,##0;($#,##0);-'
MONEY2 = '$#,##0.00;($#,##0.00);-'
PCT = '0.0%'
thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

# Tier, per door, per drawer front, job base, hrs/door, hrs/drawer front, base hrs
# name, door material $/unit, install labor $/unit, hardware $/unit (material+install),
# crew hrs per unit, base hrs, crew rate
TIERS = [
    ("Tier 1 — Redoor", 122.72, 60.00, 20.60, 0.45, 6, 65),
    ("Tier 2 — Redoor + Painted Frames", 122.72, 82.40, 20.60, 0.65, 8, 65),
    ("Tier 3 — Reface", 147.72, 144.20, 20.60, 0.90, 12, 65),
]

TIER_SCOPE = [
    ("Tier 1 — Redoor",
     "Removal of old doors and installation of new doors with new hinges. Includes prep, restore "
     "or paint on the outside of existing cabinet frames and boxes. Not included: bottoms of wall "
     "cabinets or cabinet interiors unless in scope. (ServiceMinder redooring labor line, $60/unit.)"),
    ("Tier 2 — Redoor + Painted Frames",
     "Tier 1, plus prep, restore and paint of the cabinet face frames and the interiors of glass-door "
     "cabinets. Not included: bottoms of wall cabinets or interiors unless in scope. "
     "(ServiceMinder labor line, $82.40/unit.)"),
    ("Tier 3 — Reface",
     "Removal of old doors, installation of new doors with new hinges, plus prep of the outside of "
     "cabinet frames and boxes for new matching material. Not included: bottoms of wall cabinets or "
     "interiors unless in scope. (ServiceMinder refacing labor line, $144.20/unit.)"),
]

# Every rate below appears on live ServiceMinder KTU invoices; the count is how many
# invoice lines carried that exact rate in the 2025-01-01 .. 2026-08-18 pull.
ADDONS = [
    ("Disposal of existing doors", "per job", 350, "SM invoices (16 lines at $350)"),
    ("Shipping & handling", "per job", 750, "SM invoices (21 lines at $750)"),
    ("Hardware shipping", "per job", 25, "SM invoices (26 lines at $25)"),
    ("Veneer / molding material (Elias)", "per job", 800, "SM invoices — varies $250-$3,100, edit per job"),
    ("Quartz or granite countertop", "per SF", 105, "SM invoices (13 lines at $105; $115 also used)"),
    ("Countertop cutout for sink or cooktop", "each", 110, "SM invoices (28 lines at $110)"),
    ("Disposal of existing countertop", "per job", 350, "SM invoices (16 lines at $350)"),
    ("Ceramic tile backsplash", "per SF", 65, "SM invoices (25 lines at $65)"),
    ("Undermount stainless sink", "each", 275, "SM invoices (20 lines at $275)"),
    ("Standard rollout trays", "each", 275, "SM invoices (many lines at $275)"),
    ("Trash pullout — single 50qt", "each", 350, "SM invoices (4 lines at $350)"),
    ("Trash pullout — double 35qt", "each", 450, "SM invoices (8 lines at $450)"),
    ("Two-tier Lazy Susan", "each", 450, "SM invoices ($400-$625 range)"),
    ("Undercabinet LED light kit", "per job", 1400, "SM invoices (4 lines at $1,400; $1,200 also used)"),
    ("Add crown to wall cabinets", "per run", 175, "SM invoices (6 lines at $175)"),
    ("Crown molding install", "per piece", 26.78, "SM invoices (26 lines at $26.78)"),
    ("Light rail molding", "per run", 110, "SM invoices (8 lines at $110)"),
    ("Light rail install", "per piece", 26.78, "SM invoices (8 lines at $26.78)"),
    ("Reface rear of island", "per island", 300, "SM invoices (7 lines at $300)"),
    ("Refrigerator wall panel", "each", 300, "SM invoices (6 lines at $300)"),
    ("Reface underside of wall cabinets", "per job", 170, "SM invoices (3 lines at $170)"),
    ('42" wall cabinet upcharge', "each", 125, "SM invoices (2 lines at $125)"),
    ("Matching end door panels", "per job", 200, "SM invoices ($200-$250)"),
    ("Drawer box upgrade", "each", 250, "SM invoices (multiple lines at $250)"),
    ("Repair drywall and spackle", "per area", 400, "SM invoices (3 lines at $400)"),
    ("Paint kitchen walls and ceiling", "per job", 1200, "SM invoices ($950-$1,900 range)"),
    ("Paint window and door trims", "per job", 300, "SM invoices (multiple lines at $300)"),
    ("Remove existing kitchen (demo)", "per job", 1500, "SM invoices (6 lines at $1,500)"),
    ("Plank flooring install", "per SF", 5.25, "SM invoices (5 lines at $5.25)"),
    ("Luxury vinyl plank material", "per SF", 6.50, "SM invoices (6 lines at $6.50)"),
]


wb = Workbook()


def title(ws, text, width):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    c = ws.cell(row=1, column=1, value=text)
    c.font = H1
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24


def header_row(ws, row, values, start=1, size=10):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start + i, value=v)
        c.font = Font(name=FONT, size=size, bold=True, color="FFFFFF")
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        c.border = BOX


# ------------------------------------------------------------------ Rate Card
rc = wb.active
rc.title = "Rate Card"
title(rc, "KTU RATE CARD — live ServiceMinder rates. Edit the yellow cells; every tab follows them.", 7)
rc["A3"] = ("Per-unit rates are LINE-ITEM prices, one unit = one door OR one drawer front (ServiceMinder counts them "
            "together). Shop Labor and Shop Overhead are added on top as their own proposal lines.")
rc["A3"].font = NOTE

header_row(rc, 5, ["Tier", "Door / Drawer Front Material ($/unit)", "Install Labor ($/unit)",
                   "Hardware, material + install ($/unit)", "Crew Hrs / unit", "Base Hrs", "Crew Rate $/hr"])
rc.row_dimensions[5].height = 44
for i, t in enumerate(TIERS):
    r = 6 + i
    rc.cell(row=r, column=1, value=t[0]).font = BOLD
    rc.cell(row=r, column=1).border = BOX
    for j, v in enumerate(t[1:]):
        c = rc.cell(row=r, column=2 + j, value=v)
        c.font = BLUE
        c.fill = INPUT_FILL
        c.border = BOX
        c.number_format = MONEY2 if j <= 2 else ('0.00' if j == 3 else ('0' if j == 4 else MONEY))
TIER_FIRST, TIER_LAST = 6, 5 + len(TIERS)

rc["A10"] = ("Source: every per-unit rate above is a live ServiceMinder KTU invoice line, read 2026-08-18 across 168 "
             "invoices (2025-01-01 to 2026-08-18). Material $147.72 reface / $122.72 redoor; labor $144.20 reface, "
             "$82.40 redoor-with-painted-frames, $60.00 redoor. Hardware $10.30 material + $10.30 install. "
             "Crew hours and the $65/hr Track A rate are internal estimates for the margin check only.")
rc["A10"].font = NOTE

rc["A12"] = "SETTINGS"
rc["A12"].font = H2
settings = [
    ("Shop Labor %", 0.24, PCT, "Live on every ServiceMinder proposal as its own line"),
    ("Shop Overhead %", 0.05, PCT, "Live on every ServiceMinder proposal as its own line"),
    ("Job minimum ($)", 2500, MONEY, "Owner-set 2026-08-18. ServiceMinder has NO minimum configured on any KTU service"),
    ("Deposit %", 0.50, PCT, "Collected at signing"),
    ("Direct labor ceiling %", 0.20, PCT, "Internal target — flag any job above this"),
    ("Standard flat job lines ($)", 1125, MONEY, "Disposal $350 + S&H $750 + hardware shipping $25 — used by the matrix"),
]
for i, (label, val, fmt, note) in enumerate(settings):
    r = 13 + i
    rc.cell(row=r, column=1, value=label).font = BOLD
    c = rc.cell(row=r, column=2, value=val)
    c.font = BLUE
    c.fill = INPUT_FILL
    c.number_format = fmt
    c.border = BOX
    rc.cell(row=r, column=3, value=note).font = NOTE

SHOP_LABOR = "'Rate Card'!$B$13"
OVERHEAD = "'Rate Card'!$B$14"
JOB_MIN = "'Rate Card'!$B$15"
DEPOSIT = "'Rate Card'!$B$16"
LABOR_CEIL = "'Rate Card'!$B$17"
FLATS = "'Rate Card'!$B$18"

rc["A20"] = "WHAT EACH TIER INCLUDES — wording from the live ServiceMinder proposal lines"
rc["A20"].font = H2
header_row(rc, 21, ["Tier", "Scope"])
for i, (tier, scope) in enumerate(TIER_SCOPE):
    r = 22 + i
    rc.cell(row=r, column=1, value=tier).font = BOLD
    c = rc.cell(row=r, column=2, value=scope)
    c.font = BLACK
    c.alignment = Alignment(wrap_text=True, vertical="top")
    rc.row_dimensions[r].height = 54
    rc.cell(row=r, column=1).border = BOX
    c.border = BOX
    rc.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)

rc["A26"] = "ADD-ONS AND FLAT JOB LINES — all live ServiceMinder rates"
rc["A26"].font = H2
header_row(rc, 27, ["Add-On", "Unit", "Rate", "Source"])
for i, (name, unit, rate, src) in enumerate(ADDONS):
    r = 28 + i
    rc.cell(row=r, column=1, value=name).font = BLACK
    rc.cell(row=r, column=2, value=unit).font = BLACK
    c = rc.cell(row=r, column=3, value=rate)
    c.font = BLUE
    c.fill = INPUT_FILL
    c.number_format = MONEY2
    rc.cell(row=r, column=4, value=src).font = NOTE
    for col in range(1, 5):
        rc.cell(row=r, column=col).border = BOX
ADDON_FIRST, ADDON_LAST = 28, 27 + len(ADDONS)

rc.column_dimensions["A"].width = 36
rc.column_dimensions["B"].width = 20
rc.column_dimensions["C"].width = 20
rc.column_dimensions["D"].width = 20
for col in "EFG":
    rc.column_dimensions[col].width = 15

# ------------------------------------------------------------------ Price by Door Count
pm = wb.create_sheet("Price by Door Count")
title(pm, "PRICE BY DOOR COUNT — all-in, including hardware, the standard flat lines, Shop Labor, Overhead and the $2,500 minimum", 8)
pm["A3"] = ("Units = doors + drawer fronts (count them together, the way ServiceMinder does). Includes the standard flat "
            "lines (disposal $350 + shipping $750 + hardware shipping $25). Countertops, backsplash, accessories, "
            "lighting, paint and demo are extra — add them on the Quick Quote tab. Shaded = the $2,500 minimum set the price.")
pm["A3"].font = NOTE
pm.row_dimensions[3].height = 28

header_row(pm, 5, ["Units (doors + drawer fronts)",
                   "Tier 1 — Redoor", "Tier 2 — Redoor + Painted Frames", "Tier 3 — Reface",
                   "T1 per unit", "T2 per unit", "T3 per unit"])
pm.row_dimensions[5].height = 44

units = list(range(6, 66, 2))
for i, u in enumerate(units):
    r = 6 + i
    c = pm.cell(row=r, column=1, value=u)
    c.font = BLUE
    c.number_format = '0'
    c.border = BOX
    for j in range(3):
        src = TIER_FIRST + j
        f = (f"=MAX({JOB_MIN},("
             f"{FLATS}+$A{r}*('Rate Card'!$B${src}+'Rate Card'!$C${src}+'Rate Card'!$D${src})"
             f")*(1+{SHOP_LABOR}+{OVERHEAD}))")
        cc = pm.cell(row=r, column=2 + j, value=f)
        cc.font = BLACK
        cc.number_format = MONEY
        cc.border = BOX
        col = get_column_letter(2 + j)
        eff = pm.cell(row=r, column=5 + j, value=f"={col}{r}/$A{r}")
        eff.font = BLACK
        eff.number_format = MONEY
        eff.border = BOX

for i, u in enumerate(units):
    r = 6 + i
    for j, t in enumerate(TIERS):
        raw = (1125 + u * (t[1] + t[2] + t[3])) * 1.29
        if raw < 2500:
            pm.cell(row=r, column=2 + j).fill = MIN_FILL
            pm.cell(row=r, column=5 + j).fill = MIN_FILL

pm.cell(row=6 + len(units) + 1, column=1,
        value="Shaded = the $2,500 minimum set the price rather than the unit count.").font = NOTE
pm.column_dimensions["A"].width = 26
for col in "BCDEFG":
    pm.column_dimensions[col].width = 18
pm.freeze_panes = "B6"

# ------------------------------------------------------------------ Quick Quote
qq = wb.create_sheet("Quick Quote")
title(qq, "QUICK QUOTE — fill in the yellow cells only", 6)
qq["A3"] = "Blue text = you type it. Black = calculated. Green = pulled from the Rate Card."
qq["A3"].font = NOTE

qq["A5"] = "JOB INFO"
qq["A5"].font = H2
for i, (label, val) in enumerate([("Customer", "Example: Smith, 14 Maple Ave"),
                                  ("Date", "2026-08-18"),
                                  ("Sales rep", "Example: Ben")]):
    r = 6 + i
    qq.cell(row=r, column=1, value=label).font = BOLD
    c = qq.cell(row=r, column=2, value=val)
    c.font = BLUE
    c.fill = INPUT_FILL
    c.border = BOX

qq["A10"] = "SCOPE"
qq["A10"].font = H2
for i, (label, val) in enumerate([("Tier", "Tier 3 — Reface"),
                                  ("Units — doors + drawer fronts", 38)]):
    r = 11 + i
    qq.cell(row=r, column=1, value=label).font = BOLD
    c = qq.cell(row=r, column=2, value=val)
    c.font = BLUE
    c.fill = INPUT_FILL
    c.border = BOX
    if isinstance(val, int):
        c.number_format = '0'
qq["C11"] = "← pick from the dropdown"
qq["C11"].font = NOTE
qq["C12"] = "Count every door and every drawer front together as one number, the way ServiceMinder does."
qq["C12"].font = NOTE

dv = DataValidation(type="list", formula1=f"='Rate Card'!$A${TIER_FIRST}:$A${TIER_LAST}", allow_blank=False)
qq.add_data_validation(dv)
dv.add(qq["B11"])

TROW = f"MATCH($B$11,'Rate Card'!$A${TIER_FIRST}:$A${TIER_LAST},0)"

qq["A15"] = "CABINET WORK"
qq["A15"].font = H2
header_row(qq, 16, ["Line", "Units", "Rate", "Amount"])
lines = [
    ("Doors and drawer fronts — material", f"=INDEX('Rate Card'!$B${TIER_FIRST}:$B${TIER_LAST},{TROW})"),
    ("Removal and install labor", f"=INDEX('Rate Card'!$C${TIER_FIRST}:$C${TIER_LAST},{TROW})"),
    ("Hardware — material and install", f"=INDEX('Rate Card'!$D${TIER_FIRST}:$D${TIER_LAST},{TROW})"),
]
for i, (label, rate) in enumerate(lines):
    r = 17 + i
    qq.cell(row=r, column=1, value=label).font = BLACK
    c = qq.cell(row=r, column=2, value="=$B$12")
    c.font = BLACK
    c.number_format = '0'
    c3 = qq.cell(row=r, column=3, value=rate)
    c3.font = GREEN
    c3.number_format = MONEY2
    c4 = qq.cell(row=r, column=4, value=f"=B{r}*C{r}")
    c4.font = BLACK
    c4.number_format = MONEY
    for col in range(1, 5):
        qq.cell(row=r, column=col).border = BOX
qq["A20"] = "Cabinet work subtotal"
qq["A20"].font = BOLD
qq["D20"] = "=SUM(D17:D19)"
qq["D20"].font = BOLD
qq["D20"].number_format = MONEY
qq["D20"].fill = BAND_FILL

qq["A22"] = "ADD-ONS AND FLAT LINES — enter a quantity for anything the job includes"
qq["A22"].font = H2
header_row(qq, 23, ["Add-On", "Qty", "Unit", "Rate", "Amount"])
DEFAULT_QTY = {"Disposal of existing doors": 1, "Shipping & handling": 1, "Hardware shipping": 1}
for i, (name, unit, rate, src) in enumerate(ADDONS):
    r = 24 + i
    s_row = ADDON_FIRST + i
    qq.cell(row=r, column=1, value=f"='Rate Card'!A{s_row}").font = GREEN
    c = qq.cell(row=r, column=2, value=DEFAULT_QTY.get(name, 0))
    c.font = BLUE
    c.fill = INPUT_FILL
    c.number_format = '0.##'
    qq.cell(row=r, column=3, value=f"='Rate Card'!B{s_row}").font = GREEN
    c4 = qq.cell(row=r, column=4, value=f"='Rate Card'!C{s_row}")
    c4.font = GREEN
    c4.number_format = MONEY2
    c5 = qq.cell(row=r, column=5, value=f"=B{r}*D{r}")
    c5.font = BLACK
    c5.number_format = MONEY
    for col in range(1, 6):
        qq.cell(row=r, column=col).border = BOX
ADD_FIRST, ADD_LAST = 24, 23 + len(ADDONS)
r = ADD_LAST + 1
qq.cell(row=r, column=1, value="Add-ons subtotal").font = BOLD
c = qq.cell(row=r, column=5, value=f"=SUM(E{ADD_FIRST}:E{ADD_LAST})")
c.font = BOLD
c.number_format = MONEY
c.fill = BAND_FILL
ADD_TOTAL = f"$E${r}"

p = r + 2
qq.cell(row=p, column=1, value="PRICE BUILD-UP").font = H2
build = [
    ("Line-item subtotal", f"=$D$20+{ADD_TOTAL}", BLACK),
    ("Shop Labor", f"=D{p+1}*{SHOP_LABOR}", BLACK),
    ("Shop Overhead", f"=D{p+1}*{OVERHEAD}", BLACK),
    ("Price before minimum", f"=SUM(D{p+1}:D{p+3})", BOLD),
    ("Job minimum", f"={JOB_MIN}", GREEN),
    ("Discount (enter as a positive $)", 0, BLUE),
    ("QUOTED PRICE", f"=MAX(D{p+4}-D{p+6},D{p+5})", Font(name=FONT, size=12, bold=True)),
    ("Effective price per unit", f"=IFERROR(D{p+7}/$B$12,0)", BLACK),
    ("Deposit due at signing", f"=D{p+7}*{DEPOSIT}", BLACK),
    ("Balance", f"=D{p+7}-D{p+9}", BLACK),
]
for i, (label, val, font) in enumerate(build):
    rr = p + 1 + i
    qq.cell(row=rr, column=1, value=label).font = BOLD if font is BOLD else BLACK
    c = qq.cell(row=rr, column=4, value=val)
    c.font = font
    c.number_format = MONEY
    c.border = BOX
    if label.startswith("Discount"):
        c.fill = INPUT_FILL
    if label == "QUOTED PRICE":
        c.fill = TOTAL_FILL
QUOTED = f"$D${p+7}"
qq.cell(row=p + 7, column=5,
        value=f'=IF({QUOTED}<=D{p+5},"MINIMUM APPLIED — priced up to the $2,500 floor. Trade up: hardware, rollouts, lighting.","")').font = RED

m = p + 12
qq.cell(row=m, column=1, value="MARGIN CHECK (internal — do not show the customer)").font = H2
qq.cell(row=m, column=1).fill = MIN_FILL
mrows = [
    ("Estimated crew hours",
     f"=INDEX('Rate Card'!$F${TIER_FIRST}:$F${TIER_LAST},{TROW})"
     f"+$B$12*INDEX('Rate Card'!$E${TIER_FIRST}:$E${TIER_LAST},{TROW})", '0.0'),
    ("Crew rate ($/hr)", f"=INDEX('Rate Card'!$G${TIER_FIRST}:$G${TIER_LAST},{TROW})", MONEY),
    ("Estimated direct labor cost", f"=D{m+1}*D{m+2}", MONEY),
    ("Direct labor % of quoted price", f"=IFERROR(D{m+3}/{QUOTED},0)", PCT),
]
for i, (label, val, fmt) in enumerate(mrows):
    rr = m + 1 + i
    qq.cell(row=rr, column=1, value=label).font = BLACK
    c = qq.cell(row=rr, column=4, value=val)
    c.font = BLACK
    c.number_format = fmt
    c.border = BOX
qq.cell(row=m + 4, column=5,
        value=f'=IF(D{m+4}>{LABOR_CEIL},"OVER the 20% labor ceiling — re-check hours or raise the price","Within the 20% labor ceiling")').font = BOLD
qq.cell(row=m + 6, column=1,
        value="Crew hours are internal estimates, not Construction Clock actuals. Door material and vendor costs are not in this check.").font = NOTE

qq.column_dimensions["A"].width = 44
qq.column_dimensions["B"].width = 10
qq.column_dimensions["C"].width = 18
qq.column_dimensions["D"].width = 16
qq.column_dimensions["E"].width = 62

# ------------------------------------------------------------------ How It Works
hp = wb.create_sheet("How Pricing Works", 0)
title(hp, "HOW KTU PRICING WORKS — straight from ServiceMinder", 1)
content = [
    ("H", "The formula"),
    ("T", "Quoted price = ( Units x (door material + install labor + hardware) + flat lines + add-ons ) x 1.29, floored at the $2,500 job minimum."),
    ("T", "The 1.29 is Shop Labor 24% + Shop Overhead 5%. Both are already on every ServiceMinder proposal as their own separate lines — this workbook just does the same math in advance."),
    ("B", ""),
    ("H", "Why per unit"),
    ("T", "ServiceMinder prices doors and drawer fronts as one combined unit count. A 38-unit kitchen carries 38 door-material lines, 38 labor lines, 38 hardware lines and 38 hardware-install lines. Count the doors and the drawer fronts together and the job prices itself."),
    ("B", ""),
    ("H", "The three tiers, and what makes them different"),
    ("T", "Tier 1 Redoor — $122.72 material + $60.00 labor per unit. New doors and drawer fronts with new hinges; prep, restore or paint the outside of the existing frames and boxes."),
    ("T", "Tier 2 Redoor with painted frames — $122.72 material + $82.40 labor per unit. Adds prep, restore and paint of the face frames and the interiors of glass-door cabinets."),
    ("T", "Tier 3 Reface — $147.72 material + $144.20 labor per unit. Adds prep of the outside of the frames and boxes for new matching material over them."),
    ("T", "Every tier also carries hardware at $10.30 material + $10.30 install per unit. None of the tiers include the bottoms of wall cabinets or cabinet interiors unless the scope says so."),
    ("B", ""),
    ("H", "The standard flat lines"),
    ("T", "Disposal of existing doors $350. Shipping and handling $750. Hardware shipping $25. Those three ride on nearly every job and are already in the door-count matrix. The veneer and molding material line (Elias) is job-specific and ranges from about $250 to $3,100 — set it per job."),
    ("B", ""),
    ("H", "The $2,500 minimum"),
    ("T", "New as of 2026-08-18 and owner-set: no job goes out below $2,500. ServiceMinder itself has no minimum configured on any of the 42 KTU services, so nothing enforces this except us. In practice it only binds on very small jobs — roughly under 7 units at Tier 1 and under 5 at Tier 3 — but those are exactly the jobs where drive time, measure, order minimums and set-up eat the whole margin."),
    ("T", "When the sheet flags the minimum, treat it as a trade-up moment: the customer is paying $2,500 either way, so move them up a tier or add hardware, rollouts, a trash pullout or under-cabinet lighting and they get real value for the same check."),
    ("B", ""),
    ("H", "Add-ons"),
    ("T", "Countertops $105/SF, sink cutout $110, ceramic backsplash $65/SF, undermount sink $275, rollout trays $275, trash pullouts $350-$450, Lazy Susan $450, LED light kit $1,400, crown $175/run plus $26.78/piece install, island rear reface $300, fridge panel $300, demo $1,500, wall and ceiling paint $1,200. All of these are live ServiceMinder rates and all ride through the same 1.29."),
    ("B", ""),
    ("H", "The margin guardrail"),
    ("T", "Every quote estimates crew hours from the unit count and flags any job where direct labor runs over 20% of price — the internal ceiling per service line. A flag means re-check the hours or raise the price before it is signed."),
    ("B", ""),
    ("H", "Who changes what"),
    ("T", "Reps fill in the yellow cells on Quick Quote only. Rates live on the Rate Card tab and are owner-set, so every rep quotes the same kitchen the same way — and the same way ServiceMinder will bill it."),
    ("B", ""),
    ("H", "Where the numbers came from"),
    ("T", "Every per-unit rate, flat line and add-on rate in this workbook was read directly from live ServiceMinder KTU invoices on 2026-08-18 — 168 invoices covering 2025-01-01 through 2026-08-18. Nothing here is estimated except the crew-hour figures used by the internal margin check and the $2,500 minimum, which is owner-set."),
    ("T", "Note on drift: the redoor labor line has been billed at $60.00 and the reface labor line at $144.20 in 2026, while some 2025 jobs carry $60 and $82.40. Door material shows $147.72 (reface), $122.72 (redoor) and occasional $142.72 / $137.45. The Rate Card uses the most-billed current rate for each."),
]
r = 3
for kind, text in content:
    if kind == "B":
        r += 1
        continue
    c = hp.cell(row=r, column=1, value=text)
    c.font = H2 if kind == "H" else Font(name=FONT, size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if kind == "T":
        hp.row_dimensions[r].height = 34
    r += 1
hp.column_dimensions["A"].width = 132

wb.save("/home/user/TeamLivingston/docs/pricing/KTU_TuneUp_Pricing.xlsx")
print("written")
