"""Build the KTU sales-rep door-count pricing workbook.

Rates flagged "internal" come from KTU_Custom_Kitchen_Price_Adjustments (Drive:
02 Sales & Pricing, 2026-06-05). Rates flagged "assumption" are starting defaults
set in this build and are meant to be tuned by the owner before the rep uses them.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FONT = "Arial"
BLUE = Font(name=FONT, size=10, color="0000FF")          # hardcoded input
BLACK = Font(name=FONT, size=10)                          # formula
GREEN = Font(name=FONT, size=10, color="008000")          # link to other sheet
BOLD = Font(name=FONT, size=10, bold=True)
H1 = Font(name=FONT, size=14, bold=True, color="FFFFFF")
H2 = Font(name=FONT, size=11, bold=True)
NOTE = Font(name=FONT, size=9, italic=True, color="595959")

HDR_FILL = PatternFill("solid", fgColor="1F3864")
BAND_FILL = PatternFill("solid", fgColor="D9E2F3")
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
TOTAL_FILL = PatternFill("solid", fgColor="E2EFDA")
FLAG_FILL = PatternFill("solid", fgColor="FCE4D6")

MONEY = '$#,##0;($#,##0);-'
MONEY2 = '$#,##0.00;($#,##0.00);-'
PCT = '0.0%'
thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

SERVICES = [
    # name, base, L1 door, L1 df, L2 door, L2 df, L3 door, L3 df, hrs/door, hrs/df, base hrs, labor rate
    ("Tune-Up / Reconditioning", 450, 45, 25, 60, 32, 75, 40, 0.35, 0.15, 4, 65),
    ("Cabinet Painting", 950, 135, 75, 160, 88, 190, 105, 0.80, 0.40, 8, 65),
    ("Redooring", 850, 185, 95, 225, 115, 275, 140, 0.45, 0.20, 6, 65),
    ("Refacing", 1200, 245, 125, 295, 150, 355, 180, 0.90, 0.40, 12, 65),
    ("Reface Plus", 1600, 295, 150, 350, 178, 415, 210, 1.10, 0.50, 16, 65),
]

ADDONS = [
    ("New cabinet — custom/replacement", "per cabinet", 1150, "internal"),
    ("Countertops — quartz/granite", "per SF", 105, "internal"),
    ("Backsplash — quartz/granite", "per SF", 115, "internal"),
    ("Backsplash — ceramic tile", "per SF", 96, "internal"),
    ("Plank flooring install", "per SF", 6, "internal"),
    ("Existing kitchen disposal / demo", "per job", 2300, "internal"),
    ("Wall removal / entryway", "per opening", 1535, "internal"),
    ("Paint kitchen walls and ceiling", "per job", 1920, "internal"),
    ("Paint kitchen trims", "per job", 650, "internal"),
    ("Repair drywall and spackle", "per area", 650, "internal"),
    ("Reface rear of island", "per island", 690, "internal"),
    ("Refrigerator panel", "each", 345, "internal"),
    ("Microwave drawer base", "each", 425, "internal"),
    ('42" wall cabinet upcharge', "each", 175, "internal"),
    ("Matching end panels", "each", 192, "internal"),
    ("Lazy Susan", "each", 460, "internal"),
    ("Rollout trays", "each", 275, "internal"),
    ("Trash pullout", "each", 450, "internal"),
    ("Add crown molding", "per run", 230, "internal"),
    ("Crown transition", "per piece", 230, "internal"),
    ("Hardware — material", "per piece", 17, "internal"),
    ("Hardware — install", "per piece", 17, "internal"),
    ("Undercabinet lighting", "per job", 1150, "internal"),
    ("Sink hook-up", "each", 385, "internal"),
    ("Sink allowance", "each", 275, "internal"),
]

wb = Workbook()


def title(ws, text, width):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    c = ws.cell(row=1, column=1, value=text)
    c.font = H1
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24


def header_row(ws, row, values, start=1):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start + i, value=v)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        c.border = BOX


# ---------------------------------------------------------------- Rate Card
rc = wb.active
rc.title = "Rate Card"
title(rc, "RATE CARD — edit the yellow cells to change every price in this workbook", 12)
rc["A3"] = "Per-door / per-drawer-front rates. These are LINE-ITEM prices; Shop Labor and Shop Overhead are added on top (see Settings below)."
rc["A3"].font = NOTE

header_row(rc, 5, ["Service", "Job Base", "L1 Door", "L1 Drawer Front", "L2 Door",
                   "L2 Drawer Front", "L3 Door", "L3 Drawer Front",
                   "Hrs / Door", "Hrs / Drawer Front", "Base Hrs", "Crew Rate $/hr"])
for i, s in enumerate(SERVICES):
    r = 6 + i
    rc.cell(row=r, column=1, value=s[0]).font = BOLD
    for j, v in enumerate(s[1:]):
        c = rc.cell(row=r, column=2 + j, value=v)
        c.font = BLUE
        c.fill = INPUT_FILL
        c.border = BOX
        c.number_format = MONEY if j <= 6 else ('0.00' if j <= 8 else ('0' if j == 9 else MONEY))
    rc.cell(row=r, column=1).border = BOX

rc["A12"] = ("Levels: L1 = Good (thermofoil / laminate doors, standard veneer). "
             "L2 = Better (painted MDF or wood shaker). L3 = Best (premium wood species, specialty color or finish).")
rc["A12"].font = NOTE
rc["A13"] = ("Source: per-door, per-drawer-front, job-base and labor-hour figures are ASSUMPTION defaults set 2026-08-18 — "
             "no per-door price list existed in the Drive or the repo. Owner to confirm before field use.")
rc["A13"].font = NOTE

rc["A16"] = "SETTINGS"
rc["A16"].font = H2
settings = [
    ("Shop Labor %", 0.24, PCT, "Internal — added as its own proposal line, same as ServiceMinder"),
    ("Shop Overhead %", 0.05, PCT, "Internal — added as its own proposal line"),
    ("Job minimum ($)", 2500, MONEY, "Owner-set 2026-08-18: no job quotes below this"),
    ("Deposit %", 0.50, PCT, "Collected at signing"),
    ("Direct labor ceiling %", 0.20, PCT, "Internal target — flag any job above this"),
    ("Drawer fronts per door (matrix est.)", 0.45, '0.00', "Used only by the Price by Door Count tab"),
]
for i, (label, val, fmt, note) in enumerate(settings):
    r = 17 + i
    rc.cell(row=r, column=1, value=label).font = BOLD
    c = rc.cell(row=r, column=2, value=val)
    c.font = BLUE
    c.fill = INPUT_FILL
    c.number_format = fmt
    c.border = BOX
    rc.cell(row=r, column=3, value=note).font = NOTE

rc["A25"] = "ADD-ON CATALOG"
rc["A25"].font = H2
header_row(rc, 26, ["Add-On", "Unit", "Rate", "Source"])
for i, (name, unit, rate, src) in enumerate(ADDONS):
    r = 27 + i
    rc.cell(row=r, column=1, value=name).font = BLACK
    rc.cell(row=r, column=2, value=unit).font = BLACK
    c = rc.cell(row=r, column=3, value=rate)
    c.font = BLUE
    c.fill = INPUT_FILL
    c.number_format = MONEY
    rc.cell(row=r, column=4, value=(
        "KTU internal rate card (Drive: 02 Sales & Pricing, KTU_Custom_Kitchen_Price_Adjustments, 2026-06-05)"
        if src == "internal" else "Assumption")).font = NOTE
    for col in range(1, 5):
        rc.cell(row=r, column=col).border = BOX

rc.column_dimensions["A"].width = 38
rc.column_dimensions["B"].width = 14
for col in "CDEFGHIJKL":
    rc.column_dimensions[col].width = 13
rc.column_dimensions["D"].width = 16

ADDON_FIRST, ADDON_LAST = 27, 26 + len(ADDONS)
SVC_FIRST, SVC_LAST = 6, 5 + len(SERVICES)
SHOP_LABOR = "'Rate Card'!$B$17"
OVERHEAD = "'Rate Card'!$B$18"
JOB_MIN = "'Rate Card'!$B$19"
DEPOSIT = "'Rate Card'!$B$20"
LABOR_CEIL = "'Rate Card'!$B$21"
DF_RATIO = "'Rate Card'!$B$22"

# ---------------------------------------------------------------- Quick Quote
qq = wb.create_sheet("Quick Quote")
title(qq, "QUICK QUOTE — price a job by door count", 6)
qq["A3"] = "Fill in the YELLOW cells only. Everything else calculates. Blue text = you type it; black = formula."
qq["A3"].font = NOTE

qq["A5"] = "JOB INFO"
qq["A5"].font = H2
info = [("Customer", "Example: Smith, 14 Maple Ave"), ("Date", "2026-08-18"), ("Sales rep", "Example: Rep name")]
for i, (label, val) in enumerate(info):
    r = 6 + i
    qq.cell(row=r, column=1, value=label).font = BOLD
    c = qq.cell(row=r, column=2, value=val)
    c.font = BLUE
    c.fill = INPUT_FILL
    c.border = BOX

qq["A10"] = "SCOPE"
qq["A10"].font = H2
scope = [("Service", "Refacing"), ("Level (1 / 2 / 3)", 2), ("Door count", 32), ("Drawer front count", 14)]
for i, (label, val) in enumerate(scope):
    r = 11 + i
    qq.cell(row=r, column=1, value=label).font = BOLD
    c = qq.cell(row=r, column=2, value=val)
    c.font = BLUE
    c.fill = INPUT_FILL
    c.border = BOX
    if isinstance(val, int):
        c.number_format = '0'
qq["C11"] = "← pick from the dropdown"
qq["C11"].font = NOTE
qq["C13"] = "Count every hinged door, including pantry and appliance-panel doors."
qq["C13"].font = NOTE
qq["C14"] = "Count every drawer front and false front."
qq["C14"].font = NOTE

dv_svc = DataValidation(type="list", formula1=f"='Rate Card'!$A${SVC_FIRST}:$A${SVC_LAST}", allow_blank=False)
qq.add_data_validation(dv_svc)
dv_svc.add(qq["B11"])
dv_lvl = DataValidation(type="list", formula1='"1,2,3"', allow_blank=False)
qq.add_data_validation(dv_lvl)
dv_lvl.add(qq["B12"])

SVC_ROW = f"MATCH($B$11,'Rate Card'!$A${SVC_FIRST}:$A${SVC_LAST},0)"
qq["A16"] = "CABINET WORK"
qq["A16"].font = H2
header_row(qq, 17, ["Line", "Qty", "Rate", "Amount"])
cab = [
    ("Job base (measure, delivery, set-up, clean-up)", "1",
     f"=INDEX('Rate Card'!$B${SVC_FIRST}:$B${SVC_LAST},{SVC_ROW})"),
    ("Doors", "=$B$13",
     f"=INDEX('Rate Card'!$C${SVC_FIRST}:$H${SVC_LAST},{SVC_ROW},$B$12*2-1)"),
    ("Drawer fronts", "=$B$14",
     f"=INDEX('Rate Card'!$C${SVC_FIRST}:$H${SVC_LAST},{SVC_ROW},$B$12*2)"),
]
for i, (label, qty, rate) in enumerate(cab):
    r = 18 + i
    qq.cell(row=r, column=1, value=label).font = BLACK
    c = qq.cell(row=r, column=2, value=(int(qty) if qty.isdigit() else qty))
    c.font = BLACK if qty.startswith("=") else BLUE
    c.number_format = '0'
    c3 = qq.cell(row=r, column=3, value=rate)
    c3.font = GREEN
    c3.number_format = MONEY
    c4 = qq.cell(row=r, column=4, value=f"=B{r}*C{r}")
    c4.font = BLACK
    c4.number_format = MONEY
    for col in range(1, 5):
        qq.cell(row=r, column=col).border = BOX
qq["A21"] = "Cabinet work subtotal"
qq["A21"].font = BOLD
qq["D21"] = "=SUM(D18:D20)"
qq["D21"].font = BOLD
qq["D21"].number_format = MONEY
qq["D21"].fill = BAND_FILL

qq["A23"] = "ADD-ONS — enter a quantity next to anything the job includes"
qq["A23"].font = H2
header_row(qq, 24, ["Add-On", "Qty", "Unit", "Rate", "Amount"])
for i in range(len(ADDONS)):
    r = 25 + i
    src = ADDON_FIRST + i
    qq.cell(row=r, column=1, value=f"='Rate Card'!A{src}").font = GREEN
    c = qq.cell(row=r, column=2, value=0)
    c.font = BLUE
    c.fill = INPUT_FILL
    c.number_format = '0.##'
    qq.cell(row=r, column=3, value=f"='Rate Card'!B{src}").font = GREEN
    c4 = qq.cell(row=r, column=4, value=f"='Rate Card'!C{src}")
    c4.font = GREEN
    c4.number_format = MONEY
    c5 = qq.cell(row=r, column=5, value=f"=B{r}*D{r}")
    c5.font = BLACK
    c5.number_format = MONEY
    for col in range(1, 6):
        qq.cell(row=r, column=col).border = BOX
ADD_FIRST, ADD_LAST = 25, 24 + len(ADDONS)
r = ADD_LAST + 1
qq.cell(row=r, column=1, value="Add-ons subtotal").font = BOLD
qq.cell(row=r, column=5, value=f"=SUM(E{ADD_FIRST}:E{ADD_LAST})").font = BOLD
qq.cell(row=r, column=5).number_format = MONEY
qq.cell(row=r, column=5).fill = BAND_FILL
ADD_TOTAL = f"$E${r}"

# Price build-up
p = r + 2
qq.cell(row=p, column=1, value="PRICE BUILD-UP").font = H2
rows = [
    ("Line-item subtotal", f"=$D$21+{ADD_TOTAL}", MONEY, BLACK),
    ("Shop Labor", f"=D{p+1}*{SHOP_LABOR}", MONEY, BLACK),
    ("Shop Overhead", f"=D{p+1}*{OVERHEAD}", MONEY, BLACK),
    ("Price before minimum", f"=SUM(D{p+1}:D{p+3})", MONEY, BOLD),
    ("Job minimum", f"={JOB_MIN}", MONEY, GREEN),
    ("Discount (enter as a positive $)", 0, MONEY, BLUE),
    ("QUOTED PRICE", f"=MAX(D{p+4}-D{p+6},D{p+5})", MONEY, Font(name=FONT, size=12, bold=True)),
    ("Deposit due at signing", f"=D{p+7}*{DEPOSIT}", MONEY, BLACK),
    ("Balance", f"=D{p+7}-D{p+8}", MONEY, BLACK),
]
for i, (label, val, fmt, font) in enumerate(rows):
    rr = p + 1 + i
    qq.cell(row=rr, column=1, value=label).font = BOLD if font is BOLD else BLACK
    c = qq.cell(row=rr, column=4, value=val)
    c.font = font
    c.number_format = fmt
    c.border = BOX
    if label == "Discount (enter as a positive $)":
        c.fill = INPUT_FILL
    if label == "QUOTED PRICE":
        c.fill = TOTAL_FILL
QUOTED = f"$D${p+7}"
qq.cell(row=p + 7, column=5,
        value=f'=IF({QUOTED}={JOB_MIN},"MINIMUM APPLIED — job priced up to the ${{}} floor","")'
        .replace("{}", "2,500")).font = Font(name=FONT, size=10, bold=True, color="C00000")

# Margin check
m = p + 11
qq.cell(row=m, column=1, value="MARGIN CHECK (internal — do not show the customer)").font = H2
qq.cell(row=m, column=1).fill = FLAG_FILL
mrows = [
    ("Estimated crew hours",
     f"=INDEX('Rate Card'!$K${SVC_FIRST}:$K${SVC_LAST},{SVC_ROW})"
     f"+$B$13*INDEX('Rate Card'!$I${SVC_FIRST}:$I${SVC_LAST},{SVC_ROW})"
     f"+$B$14*INDEX('Rate Card'!$J${SVC_FIRST}:$J${SVC_LAST},{SVC_ROW})", '0.0'),
    ("Crew rate ($/hr)", f"=INDEX('Rate Card'!$L${SVC_FIRST}:$L${SVC_LAST},{SVC_ROW})", MONEY),
    ("Estimated direct labor cost", f"=D{m+1}*D{m+2}", MONEY),
    ("Direct labor % of quoted price", f"=IFERROR(D{m+3}/{QUOTED},0)", PCT),
]
for i, (label, val, fmt) in enumerate(mrows):
    rr = m + 1 + i
    qq.cell(row=rr, column=1, value=label).font = BLACK
    c = qq.cell(row=rr, column=4, value=val)
    c.font = BLACK
    c.number_format = fmt
    c.border = BOX
qq.cell(row=m + 4, column=5,
        value=f'=IF(D{m+4}>{LABOR_CEIL},"OVER the 20% labor ceiling — re-check hours or raise the price","Within the 20% labor ceiling")').font = BOLD
qq.cell(row=m + 6, column=1,
        value="Labor hours are estimates from scope, not Construction Clock actuals. Materials, subs and vendor costs are not in this check.").font = NOTE

qq.column_dimensions["A"].width = 46
qq.column_dimensions["B"].width = 10
qq.column_dimensions["C"].width = 14
qq.column_dimensions["D"].width = 16
qq.column_dimensions["E"].width = 52

# ---------------------------------------------------------------- Price by Door Count
pm = wb.create_sheet("Price by Door Count")
title(pm, "PRICE BY DOOR COUNT — all-in quoted price, including Shop Labor, Overhead and the job minimum", 17)
pm["A3"] = ("Drawer fronts are estimated at the ratio on the Rate Card (default 0.45 per door). "
            "Add-ons are NOT included — add them from the Quick Quote tab. Shaded cells are jobs the minimum lifted.")
pm["A3"].font = NOTE

header_row(pm, 5, ["Doors", "Drawer fronts (est.)"])
col = 3
for s in SERVICES:
    for lvl in (1, 2, 3):
        c = pm.cell(row=5, column=col, value=f"{s[0]}\nL{lvl}")
        c.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        c.border = BOX
        col += 1
pm.row_dimensions[5].height = 42

doors = list(range(8, 62, 2))
for i, d in enumerate(doors):
    r = 6 + i
    c = pm.cell(row=r, column=1, value=d)
    c.font = BLUE
    c.number_format = '0'
    c.border = BOX
    c2 = pm.cell(row=r, column=2, value=f"=ROUND($A{r}*{DF_RATIO},0)")
    c2.font = BLACK
    c2.number_format = '0'
    c2.border = BOX
    col = 3
    for si in range(len(SERVICES)):
        src = SVC_FIRST + si
        for lvl in (1, 2, 3):
            door_col = get_column_letter(2 + lvl * 2 - 1)   # C, E, G
            df_col = get_column_letter(2 + lvl * 2)         # D, F, H
            base = f"'Rate Card'!$B${src}"
            f = (f"=MAX({JOB_MIN},("
                 f"{base}+$A{r}*'Rate Card'!${door_col}${src}+$B{r}*'Rate Card'!${df_col}${src}"
                 f")*(1+{SHOP_LABOR}+{OVERHEAD}))")
            cc = pm.cell(row=r, column=col, value=f)
            cc.font = BLACK
            cc.number_format = MONEY
            cc.border = BOX
            col += 1

pm.column_dimensions["A"].width = 8
pm.column_dimensions["B"].width = 12
for i in range(3, 18):
    pm.column_dimensions[get_column_letter(i)].width = 13
pm.freeze_panes = "C6"

# ---------------------------------------------------------------- How Pricing Works
hp = wb.create_sheet("How Pricing Works", 0)
title(hp, "HOW KTU TUNE-UP PRICING WORKS", 2)
lines = [
    ("H", "The formula"),
    ("T", "Quoted price = ( Job base + Doors x Door rate + Drawer fronts x Drawer-front rate + Add-ons ) x 1.29, floored at the $2,500 job minimum."),
    ("T", "The 1.29 is Shop Labor 24% + Shop Overhead 5%, which go on the proposal as their own two lines exactly as they do in ServiceMinder."),
    ("B", ""),
    ("H", "Why door count"),
    ("T", "On every tune-up service the cabinet boxes stay put, so the work — and the material — scales with the number of doors and drawer fronts, not with the square footage of the kitchen. Count the doors and the drawer fronts and the job prices itself."),
    ("B", ""),
    ("H", "The five services"),
    ("T", "Tune-Up / Reconditioning — restore the existing finish. 1-2 days. No color or style change; not for painted or laminate doors."),
    ("T", "Cabinet Painting — doors and fronts sprayed off-site, boxes and face frames done on-site. About 5 days."),
    ("T", "Redooring — new doors and drawer fronts, existing boxes and existing box finish stay. 1-2 days on site."),
    ("T", "Refacing — new doors and drawer fronts plus matching veneer over the boxes. 2-4 days on site."),
    ("T", "Reface Plus — refacing plus new panels, fillers and trim so the boxes read as new cabinetry."),
    ("B", ""),
    ("H", "The three levels"),
    ("T", "L1 Good — thermofoil or laminate doors, standard veneer."),
    ("T", "L2 Better — painted MDF or wood shaker."),
    ("T", "L3 Best — premium wood species, specialty color or finish."),
    ("B", ""),
    ("H", "The $2,500 job minimum"),
    ("T", "No job goes out below $2,500. Below that number the drive time, measure, order minimums and set-up eat the margin. If the math lands under $2,500 the Quick Quote tab prices the job up to $2,500 and flags it. Use it as a trade-up moment: at that size the customer can usually add hardware, rollouts or lighting and get real value for the same check."),
    ("B", ""),
    ("H", "Add-ons"),
    ("T", "Countertops, backsplash, flooring, hardware, crown, lighting, accessories, demo and paint all price separately, per the add-on catalog on the Rate Card tab. They ride through the same Shop Labor and Overhead math."),
    ("B", ""),
    ("H", "The margin guardrail"),
    ("T", "Each quote estimates crew hours from the scope and flags any job where direct labor runs over 20% of price. That 20% is the internal ceiling per service line. A flag is not a stop sign — it means re-check the hour estimate or raise the price before it is signed."),
    ("B", ""),
    ("H", "What the reps may change"),
    ("T", "Reps fill in the yellow cells on Quick Quote only: customer, scope, quantities, discount. Rates live on the Rate Card tab and are changed by the owner, so every rep quotes the same job the same way."),
    ("B", ""),
    ("H", "Where the numbers came from"),
    ("T", "Add-on rates are the KTU internal rate card (Drive: 02 Sales & Pricing / KTU_Custom_Kitchen_Price_Adjustments, 2026-06-05). Shop Labor 24% and Shop Overhead 5% are the standing internal proposal structure. The 20% labor ceiling and the $65/hr reface crew rate are internal targets."),
    ("T", "The per-door, per-drawer-front and job-base rates are ASSUMPTION defaults set on 2026-08-18 — no per-door price list existed in the Drive or in the repo. Confirm them against a few recent signed reface and redoor proposals before the rep quotes live work."),
]
r = 3
for kind, text in lines:
    if kind == "B":
        r += 1
        continue
    c = hp.cell(row=r, column=1, value=text)
    c.font = H2 if kind == "H" else Font(name=FONT, size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if kind == "T":
        hp.row_dimensions[r].height = 30
    r += 1
hp.column_dimensions["A"].width = 130

wb.save("/home/user/TeamLivingston/docs/pricing/KTU_Tune-Up_Pricing_Calculator.xlsx")
print("written")
